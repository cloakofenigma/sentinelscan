"""
Excel Report Generator for SentinelScan

Generates detailed Excel (.xlsx) reports with styled columns for findings.
"""

from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..models import Finding, ScanResult
from .impact import get_impact


# Column definitions
COLUMNS = [
    ("Finding ID", 15),
    ("Title", 35),
    ("OWASP Category", 18),
    ("Severity", 12),
    ("CWE", 12),
    ("Location", 60),
    ("Description", 50),
    ("Impact", 55),
    ("Mitigation", 55),
]

# Severity fill colors
SEVERITY_FILLS = {
    "critical": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid") if HAS_OPENPYXL else None,
    "high": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid") if HAS_OPENPYXL else None,
    "medium": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid") if HAS_OPENPYXL else None,
    "low": PatternFill(start_color="28A745", end_color="28A745", fill_type="solid") if HAS_OPENPYXL else None,
    "info": PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid") if HAS_OPENPYXL else None,
}

# Severity font colors (white text for dark backgrounds, black for light)
SEVERITY_FONTS = {
    "critical": Font(color="FFFFFF", bold=True) if HAS_OPENPYXL else None,
    "high": Font(color="000000", bold=True) if HAS_OPENPYXL else None,
    "medium": Font(color="000000", bold=True) if HAS_OPENPYXL else None,
    "low": Font(color="FFFFFF", bold=True) if HAS_OPENPYXL else None,
    "info": Font(color="FFFFFF", bold=True) if HAS_OPENPYXL else None,
}


def generate_excel_report(
    scan_result: ScanResult,
    output_path: str,
    base_path: Optional[str] = None,
) -> str:
    """
    Generate Excel report from scan results.

    Args:
        scan_result: Scan results
        output_path: Path to write .xlsx file
        base_path: Base path (not stripped from location - absolute paths used)

    Returns:
        Output file path
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel reports. Install with: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Security Findings"

    # --- Header row styling ---
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sort findings by severity ---
    severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    sorted_findings = sorted(
        scan_result.findings,
        key=lambda f: severity_order.get(f.severity.value, 5),
    )

    # --- Write finding rows ---
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, finding in enumerate(sorted_findings, start=2):
        sev = finding.severity.value
        impact_text = get_impact(finding)
        mitigation_text = finding.remediation or "Refer to the CWE/OWASP reference for recommended remediation steps."

        row_data = [
            finding.rule_id,
            finding.rule_name,
            finding.owasp or "",
            sev.upper(),
            finding.cwe or "",
            f"{finding.location.file_path}:{finding.location.line_number}",
            finding.description,
            impact_text,
            mitigation_text,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Style the severity column
            if col_idx == 4:
                fill = SEVERITY_FILLS.get(sev)
                font = SEVERITY_FONTS.get(sev)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # --- Add summary sheet ---
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    summary_header_font = Font(bold=True, size=14, color="0F3460")
    ws_summary.cell(row=1, column=1, value="SentinelScan Report Summary").font = summary_header_font
    ws_summary.merge_cells("A1:B1")

    ws_summary.cell(row=3, column=1, value="Total Findings").font = Font(bold=True)
    ws_summary.cell(row=3, column=2, value=len(scan_result.findings))

    ws_summary.cell(row=4, column=1, value="Files Scanned").font = Font(bold=True)
    ws_summary.cell(row=4, column=2, value=scan_result.files_scanned)

    ws_summary.cell(row=4, column=1, value="Files Scanned").font = Font(bold=True)
    ws_summary.cell(row=4, column=2, value=scan_result.files_scanned)

    ws_summary.cell(row=6, column=1, value="Severity").font = Font(bold=True)
    ws_summary.cell(row=6, column=2, value="Count").font = Font(bold=True)
    for cell in ws_summary[6]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

    summary = scan_result.summary
    for i, sev in enumerate(["critical", "high", "medium", "low", "info"]):
        row = 7 + i
        count = summary.get(sev, 0)
        cell_a = ws_summary.cell(row=row, column=1, value=sev.upper())
        cell_b = ws_summary.cell(row=row, column=2, value=count)
        cell_a.border = thin_border
        cell_b.border = thin_border
        fill = SEVERITY_FILLS.get(sev)
        font = SEVERITY_FONTS.get(sev)
        if fill:
            cell_a.fill = fill
        if font:
            cell_a.font = font

    # Auto-filter on findings sheet
    ws_findings = wb["Security Findings"]
    ws_findings.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(sorted_findings) + 1}"

    wb.save(output_path)
    return output_path


class ExcelReporter:
    """Excel reporter class."""

    def write(
        self,
        scan_result: ScanResult,
        output_path: str,
        base_path: Optional[str] = None,
    ) -> str:
        """Write Excel report to file."""
        return generate_excel_report(scan_result, output_path, base_path)
