"""Tests for scanengine.reporters (impact, SARIF, HTML, Excel)"""

import json
import pytest
from pathlib import Path
from scanengine.models import Finding, ScanResult, Severity, Confidence, Location
from scanengine.reporters.impact import get_impact, CWE_IMPACT, TAG_IMPACT, SEVERITY_IMPACT
from scanengine.reporters.sarif import generate_sarif, SARIFReporter, write_sarif
from scanengine.reporters.html import generate_html_report, HTMLReporter, escape_html
from scanengine.reporters.excel import ExcelReporter


# ── Impact Tests ──

class TestImpact:
    def test_cwe_lookup(self, sample_finding):
        # sample_finding has CWE-89
        impact = get_impact(sample_finding)
        assert "SQL" in impact or "database" in impact

    def test_tag_fallback(self):
        finding = Finding(
            rule_id="X-001", rule_name="Test", description="d",
            severity=Severity.HIGH, confidence=Confidence.HIGH,
            location=Location(file_path="f.py", line_number=1),
            tags=["command-injection"],
        )
        impact = get_impact(finding)
        assert "command" in impact.lower() or "system" in impact.lower()

    def test_severity_fallback(self):
        finding = Finding(
            rule_id="X-001", rule_name="Test", description="d",
            severity=Severity.MEDIUM, confidence=Confidence.LOW,
            location=Location(file_path="f.py", line_number=1),
        )
        impact = get_impact(finding)
        assert impact == SEVERITY_IMPACT["medium"]

    def test_all_cwe_entries_are_strings(self):
        for key, val in CWE_IMPACT.items():
            assert isinstance(key, str)
            assert isinstance(val, str)


# ── SARIF Tests ──

class TestSARIF:
    def test_sarif_structure(self, sample_scan_result):
        sarif = generate_sarif(sample_scan_result)
        assert sarif["$schema"].endswith("sarif-schema-2.1.0.json")
        assert sarif["version"] == "2.1.0"
        assert "runs" in sarif
        assert len(sarif["runs"]) == 1

    def test_sarif_tool_info(self, sample_scan_result):
        sarif = generate_sarif(sample_scan_result)
        driver = sarif["runs"][0]["tool"]["driver"]
        assert driver["name"] == "SentinelScan"
        assert "rules" in driver

    def test_sarif_rules_deduped(self, sample_scan_result):
        sarif = generate_sarif(sample_scan_result)
        rules = sarif["runs"][0]["tool"]["driver"]["rules"]
        rule_ids = [r["id"] for r in rules]
        assert len(rule_ids) == len(set(rule_ids))

    def test_sarif_results_count(self, sample_scan_result):
        sarif = generate_sarif(sample_scan_result)
        results = sarif["runs"][0]["results"]
        assert len(results) == len(sample_scan_result.findings)

    def test_sarif_result_fields(self, sample_scan_result):
        sarif = generate_sarif(sample_scan_result)
        result = sarif["runs"][0]["results"][0]
        assert "ruleId" in result
        assert "level" in result
        assert "message" in result
        assert "locations" in result

    def test_sarif_write_file(self, sample_scan_result, tmp_path):
        output = tmp_path / "results.sarif"
        write_sarif(sample_scan_result, str(output))
        assert output.exists()
        data = json.loads(output.read_text())
        assert data["version"] == "2.1.0"

    def test_sarif_reporter_class(self, sample_scan_result):
        reporter = SARIFReporter()
        sarif = reporter.generate(sample_scan_result)
        assert "runs" in sarif


# ── HTML Tests ──

class TestHTML:
    def test_escape_html(self):
        assert escape_html("<script>alert('xss')</script>") == "&lt;script&gt;alert(&#x27;xss&#x27;)&lt;/script&gt;"

    def test_escape_html_none(self):
        assert escape_html(None) == ""
        assert escape_html("") == ""

    def test_html_report_contains_findings(self, sample_scan_result):
        html = generate_html_report(sample_scan_result)
        assert "SQLI-001" in html
        assert "CMD-001" in html
        assert "LOG-001" in html

    def test_html_report_severity_counts(self, sample_scan_result):
        html = generate_html_report(sample_scan_result)
        # Chart.js data should include severity counts
        assert "Critical" in html
        assert "High" in html

    def test_html_report_description_impact_mitigation(self, sample_scan_result):
        html = generate_html_report(sample_scan_result)
        assert "finding-detail-block description" in html
        assert "finding-detail-block impact" in html
        assert "finding-detail-block mitigation" in html

    def test_html_report_title(self, sample_scan_result):
        html = generate_html_report(sample_scan_result, title="My Scan")
        assert "My Scan" in html

    def test_html_reporter_class(self, sample_scan_result):
        reporter = HTMLReporter()
        html = reporter.generate(sample_scan_result)
        assert "<!DOCTYPE html>" in html

    def test_html_write_file(self, sample_scan_result, tmp_path):
        reporter = HTMLReporter()
        output = tmp_path / "report.html"
        reporter.write(sample_scan_result, str(output))
        assert output.exists()
        content = output.read_text()
        assert "<!DOCTYPE html>" in content


# ── Excel Tests ──

class TestExcel:
    def test_excel_creates_file(self, sample_scan_result, tmp_path):
        output = tmp_path / "report.xlsx"
        reporter = ExcelReporter()
        reporter.write(sample_scan_result, str(output))
        assert output.exists()
        assert output.stat().st_size > 0

    def test_excel_columns(self, sample_scan_result, tmp_path):
        from openpyxl import load_workbook
        output = tmp_path / "report.xlsx"
        ExcelReporter().write(sample_scan_result, str(output))

        wb = load_workbook(str(output))
        ws = wb["Security Findings"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 10)]
        assert "Finding ID" in headers
        assert "Title" in headers
        assert "Severity" in headers
        assert "CWE" in headers
        assert "Description" in headers
        assert "Impact" in headers
        assert "Mitigation" in headers

    def test_excel_row_count(self, sample_scan_result, tmp_path):
        from openpyxl import load_workbook
        output = tmp_path / "report.xlsx"
        ExcelReporter().write(sample_scan_result, str(output))

        wb = load_workbook(str(output))
        ws = wb["Security Findings"]
        # Header + 3 findings = 4 rows
        data_rows = ws.max_row - 1
        assert data_rows == len(sample_scan_result.findings)

    def test_excel_summary_sheet(self, sample_scan_result, tmp_path):
        from openpyxl import load_workbook
        output = tmp_path / "report.xlsx"
        ExcelReporter().write(sample_scan_result, str(output))

        wb = load_workbook(str(output))
        assert "Summary" in wb.sheetnames
        ws = wb["Summary"]
        assert ws.cell(row=1, column=1).value == "SentinelScan Report Summary"

    def test_excel_empty_scan(self, tmp_path):
        result = ScanResult(target_path="/empty", files_scanned=10, rules_applied=50)
        output = tmp_path / "empty.xlsx"
        ExcelReporter().write(result, str(output))
        assert output.exists()
